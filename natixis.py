from openpyxl import load_workbook
import re
import datetime
import xlsxwriter
from xlsxwriter.utility import xl_range
from xlsxwriter.utility import xl_rowcol_to_cell
import json

def format_long_sentence(label):
    new_label=label.split(' ')

    words= list()
    array = list()

    for i in range(0,len(new_label)):
        if new_label[i]!='':
                words.append(new_label[i])
        if i < len(new_label)-1:
            if new_label[i+1]=='' and len(words)>0:
                array.append(words)
                words = []
        else:
            if len(words)>0:
                array.append(words)
    clean_array = list()
    for item in array:
        clean_array.append(" ".join(item))

    regex = r"([a-zA-Z]{2}[0-9]{10})$"
    m = re.search(regex,clean_array[3])
    if m!= None:
        backup = clean_array[2]
        clean_array.pop(2)
        clean_array[1] = clean_array[1] + " " + backup
    return clean_array

# Récupère les contracts clients au sein du fichier 'Fair.xlsx' récupéré auprès
# de Fiancière Arbevel
def get_pdf_data():
    wb = load_workbook(filename='natixis.xlsx')
    ws = wb[wb.sheetnames[0]]


    regex_account_1 = r"2\.\sCompte\s(\d+)"
    regex_account_2 = r"3\.\sCompte\s(\d+)"
    regex_value = r"TOTAL\s:\s+((\S+)\sEUR)"

    account_data = list()

    liquid_start = list()
    liquid_end = list()

    begin_row = list()
    limit_account_2 = list()

    config_liquid={
        "devise":1,
        "denomination":2,
        "montant":10
    }

    for row in ws.rows:
        if row[0].value!=None:
            if type(row[0].value) is str:
                m_account_1 = re.search(regex_account_1,row[0].value)
                m_account_2 = re.search(regex_account_2,row[0].value)
                if m_account_1 != None and row[0].row > 20:
                    account_1 = m_account_1.group(1)
                    begin_row.append(row[0].row)
                if m_account_2 != None and row[0].row > 20:
                    account_2 = m_account_2.group(1)
                    begin_row.append(row[0].row)

            if row[0].value == 'Autres positions LIQUIDITÉS ET TRÉSORERIE':
                liquid_start.append(row[0].row)
            if row[0].value == "GESTION PERSONNALISÉE - ORIENTÉ ÉQUILIBRÉ":
                m_value = re.search(regex_value,ws.cell(row=row[0].row+1,column=1).value)
                print("Total Compte :",m_value.group(1))
            if type(row[0].value) is str and "3. Compte" in row[0].value:
                liquid_end.append(row[0].row)
            if type(row[0].value) is str and "4. Synthèse" in row[0].value:
                liquid_end.append(row[0].row)
            if row[0].value == "ACTIONS ET TITRES ASSIMILÉS":
                limit_account_2.append(row[0].row)
            if row[0].value == "LIQUIDITÉS ET TRÉSORERIE" and row[0].row>50:
                limit_account_2.append(row[0].row+1)
    liquid_end.pop(0)
    liquid_end.pop(0)
    print(limit_account_2)
    print("Compte",account_1)

    print('------- Liquidités ------------')

    account_data_1 = {"account_number":account_1}
    account_data_1['COMPTES COURANTS'] = list()

    position = {}
    for row in range(liquid_start[0]+3,liquid_end[0]-2):
        for key in config_liquid.keys():
            position[key] = ws.cell(row=row,column=config_liquid[key]).value
        account_data_1['COMPTES COURANTS'].append(position)
    account_data.append(account_data_1)
    print(account_data)


    account_data_2 = {"account_number":account_2}

    # Summary account 2
    key = ''
    dict_summary_2={}
    regex_subkey = r"((\S+\s)+)\s+((\S+)\sEUR)"
    for row_idx in range(limit_account_2[0],limit_account_2[1]+1):
        if ws.cell(row=row_idx,column=1).value != None:
            if ws.cell(row=row_idx,column=1).value.isupper():
                key = ws.cell(row=row_idx,column=1).value
                value = ws.cell(row=row_idx,column=12).value
                dict_summary_2[key]={
                    'total': value
                }
            else:
                m_subkey = re.search(regex_subkey,ws.cell(row=row_idx,column=1).value)
                if m_subkey != None:
                    sub_key = m_subkey.group(1)
                    value = m_subkey.group(3)
                    dict_summary_2[key][sub_key] = value
    print('---------Overview-----------')
    print(dict_summary_2)

    # Spot categories from dict_summary_2
    categories = list()
    for key in dict_summary_2.keys():
        for keyb in dict_summary_2[key].keys():
            if keyb != "total":
                account_data_2[keyb.upper().strip()]=list()
                categories.append(keyb.upper().strip())

    # Fill category positions
    category_positions = list()
    for row_idx in range(limit_account_2[1]+2,ws.max_row):
        key = ws.cell(row=row_idx,column=1).value
        if key != None and type(key) is str and len(key)>100:
            if key.strip() in categories:
                category_positions.append(row_idx)

    config_position_l={
        "devise":0,
        "denomination":1,
        "isin":2,
        "quantity": 3,
        "buy_price": 4,
        "price": 5,
        "valo_devise":6,
        "valo":7,
        "pnl":8,
        "pnl_percent":10
    }

    config_position_s={
        "devise":0,
        "denomination":1,
        "isin":2,
        "quantity": 3,
        "buy_price": 4,
        "price": 5,
        "date":6,
        "valo_devise":7,
        "valo":8,
        "pnl":9,
        "pnl_percent":10,
        "percentage_total":11
    }

    print(category_positions)


    for j in range(0,len(categories)-1):
        words=list()
        # print("-----------",categories[j],"----------")
        for row_idx in range(category_positions[j]+2,category_positions[j+1]):
            content = ws.cell(row=row_idx,column=1).value
            position = {}
            if type(content) is str:
                if len(content)>200:
                    data = format_long_sentence(content)
                    # print('long',data)
                    if len(data)>9:
                        for key in config_position_l.keys():
                            position[key]=data[config_position_l[key]]
                        position["percentage_total"] = data[9].split('\n')[0]
                        position["date"] = data[9].split('\n')[1]
                    # print(position)
                    account_data_2[categories[j]].append(position)
                else:
                    words.append(content)
            elif type(content) is datetime.datetime:
                words.append(content.strftime("%d.%m.%Y"))

            if len(words)==6:
                position={}
                data = format_long_sentence('    '.join(words))
                # print("short",data)
                for key in config_position_s.keys():
                    position[key]=data[config_position_s[key]]
                account_data_2[categories[j]].append(position)
    account_data.append(account_data_2)
    print(account_data_2)

    return account_data

def create_excel(data):
    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook('natixis_clean.xlsx')
    worksheet = workbook.add_worksheet()
    # worksheet.hide_gridlines(2)

    # Add a bold format to use to highlight cells.
    bold = workbook.add_format({'bold': True})

    # Create a format to use in the merged range.
    merge_format = workbook.add_format({
        'bold': 1,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter'})

    left_right = workbook.add_format({
        'right': True,
        'left': True
        })

    left = workbook.add_format({
        'left': True,
        })

    top = workbook.add_format({
        'top': True,
        })

    # Add a number format for cells with money.
    percent = workbook.add_format({'num_format': '0,00 %'})

    account=data[1]
    header=[
        "devise",
        "denomination",
        "isin",
        "quantity",
        "buy_price",
        "price",
        "date",
        "valo_devise",
        "valo",
        "pnl",
        "pnl_percent",
        "percentage_total"
    ]

    # Write some data headers.
    worksheet.merge_range('A1:M1','COMPTE n°'+account["account_number"],merge_format)
    worksheet.write('A2', 'categorie', merge_format)
    col = 0
    for key in header:
        worksheet.write(1,col+1, key, merge_format)
        col=col+1

    row = 2
    for key in data[1].keys():
        if key!="account_number":
            category = key
            worksheet.write(row,0, category)
            for position in data[1][category]:
                col = 1
                for keyp in header:
                    if keyp in ['pnl_percent','percentage_total']:
                        format=percent
                    else:
                        format=''
                    worksheet.write(row,col, position[keyp],format)
                    col=col+1
                row=row+1
    # for c in clients_dict.keys():
    #     if c != "_id":
    #         account_id = 1
    #         for account in clients_dict[c]:
    #             if account_id == 1:
    #                 worksheet.write(row,col, c,bold)
    #             else:
    #                 worksheet.write(row,col, "",left_right)
    #             worksheet.write(row,col+1, account_id,left_right)
    #             worksheet.write(row,col+2, account['host'],left_right)
    #             worksheet.write(row,col+3, account['type'],left_right)
    #             if 'sub_section' in account.keys():
    #                 worksheet.write(row,col+4, account['sub_section'],left_right)
    #             else:
    #                 worksheet.write(row,col+4, "",left_right)
    #             worksheet.write(row,col+5, str(account['account_number']),left_right)
    #             row = row+1
    #             account_id = account_id+1
    workbook.close()

data = get_pdf_data()
create_excel(data)
